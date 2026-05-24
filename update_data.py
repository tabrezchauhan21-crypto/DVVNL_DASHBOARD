"""
update_data.py v5 — DVVNL Dashboard Auto-Update
Extracts all 4 sheets from Invoice_Verification_Data_Status.xlsx:
  Form            → 10KW+ verification records (MONTHLY_STATUS)
  5KW_INVOICE_DATA → 5KW–9KW financial records (RECORDS_5KW)
  10KW_DATA       → 10KW+ monthly payment status (RECORDS_10KW)
  CLUSTER-4_DATA  → Cluster-4 billing financial records (RECORDS_C4)
Outputs: data_v4.js + index.html (via template)
Runs: GitHub Actions every hour via SP_5KW secret (OneDrive link)
"""
import os, json, re, io
from datetime import datetime

MONTHS = ['January','February','March','April','May','June',
          'July','August','September','October','November','December']

ZONE_MAP = {
    'AGRA':'AGRA',        'AGRA ZONE':'AGRA',
    'ALIGARH':'ALIGARH',  'ALIGARH ZONE':'ALIGARH',
    'BANDA':'BANDA',       'BANDA ZONE':'BANDA',
    'ETAH':'ETAH',         'ETAH ZONE':'ETAH',
    'FIROZABAD':'FIROZABAD','FIROZABAD ZONE':'FIROZABAD',
    'JHANSI':'JHANSI',     'JHANSI ZONE':'JHANSI',
    'KANPUR ZONE-1':'KANPUR-1','KANPUR-1':'KANPUR-1',
    'KANPUR ZONE-2':'KANPUR-2','KANPUR-2':'KANPUR-2',
    'MATHURA':'MATHURA',   'MATHURA ZONE':'MATHURA',
}

SKIP_ZONES = {'rate','grand total','total','zone_name','survey_nov-2023',''}

def sf(v):
    try: return round(float(v), 2)
    except: return 0.0

def zone(raw):
    r = str(raw or '').strip()
    return ZONE_MAP.get(r.upper(), ZONE_MAP.get(r, ''))

print("=" * 55)
print("  DVVNL Dashboard — Auto Update v5")
print("=" * 55)

# ── 1. Download Excel ─────────────────────────────────────
import openpyxl, requests

SP_LINK = os.environ.get('SP_5KW', '')
wb = None

if SP_LINK:
    def get_urls(link):
        urls = [link + ('&' if '?' in link else '?') + 'download=1']
        m = re.search(r'/g/personal/([^/]+)/([A-Za-z0-9_-]+)\?', link)
        if m:
            user, fid = m.group(1), m.group(2)
            urls.append(f"https://itsaicomputers-my.sharepoint.com/personal/{user}/_layouts/15/download.aspx?share={fid}")
        urls.append("https://itsaicomputers-my.sharepoint.com/personal/tabrez_alam_thesaicomputers_com/_layouts/15/download.aspx?UniqueId=E56B3B6C-EDB2-4425-966F-DCAB46B35F4B")
        return urls

    hdrs = {'User-Agent':'Mozilla/5.0','Accept':'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*'}
    print(f"\n[Excel] Downloading from OneDrive...")
    for i, url in enumerate(get_urls(SP_LINK)):
        try:
            r = requests.get(url, timeout=60, headers=hdrs, allow_redirects=True)
            print(f"[Excel] URL {i+1}: {len(r.content)//1024} KB | {r.headers.get('Content-Type','?')[:40]}")
            if len(r.content) < 5000: continue
            wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
            print(f"[Excel] ✅ Sheets: {wb.sheetnames}")
            break
        except Exception as e:
            print(f"[Excel] URL {i+1} failed: {e}")
else:
    print("[Excel] SP_5KW not set — trying local file...")

if not wb:
    for fname in ['Invoice_Verification_Data_Status.xlsx', 'Invoice_Verification_Data_Statusw.xlsx']:
        try:
            wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
            print(f"[Excel] ✅ Local file loaded: {fname} | Sheets: {wb.sheetnames}")
            break
        except: pass

if not wb:
    print("ERROR: Cannot load Excel. Exiting.")
    exit(1)

# ── 2. Form → 10KW verification (MONTHLY_STATUS) ─────────
print("\n[Form] Reading 10KW+ verification data...")
form_records = []
ws_form = wb['Form']
for row in ws_form.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    mon = str(row[6] or '').strip()
    yr  = str(row[7] or '').strip()
    if not mon or not yr: continue
    z = zone(row[10])
    if not z: z = str(row[10] or '').strip()   # fallback: keep raw
    form_records.append({
        'zone':        z,
        'verified_by': str(row[11] or '').strip(),
        'test_div':    str(row[12] or '').strip(),
        'dist_div':    str(row[13] or '').strip(),
        'circle':      str(row[14] or '').strip(),
        'month':       mon,
        'year':        yr,
        'letter_no':   str(row[9]  or '').strip(),
        'letter_date': row[8].strftime('%Y-%m-%d') if isinstance(row[8], datetime) else '',
    })
print(f"[Form] Records: {len(form_records)}")

# ── 3. Master data → division list ───────────────────────
print("\n[Master] Loading master_data.json...")
with open('master_data.json') as f:
    master = json.load(f)

master_dist, master_test, master_circle = [], [], []
seen_t, seen_c = set(), set()
for m in master:
    master_dist.append({'zone':m['zone'],'circle':m['circle'],'division':m['dist'],'type':'DIST'})
    if m['test'] not in seen_t:
        seen_t.add(m['test'])
        master_test.append({'zone':m['zone'],'circle':m['circle'],'division':m['test'],'type':'TEST'})
    if m['circle'] not in seen_c:
        seen_c.add(m['circle'])
        master_circle.append({'zone':m['zone'],'circle':m['circle'],'division':m['circle'],'type':'CIRCLE'})
MASTER_ALL = master_dist + master_test + master_circle
print(f"[Master] Divisions: {len(MASTER_ALL)}")

# ── 4. Compute MONTHLY_STATUS ─────────────────────────────
print("\n[Form] Computing monthly verification status...")
all_months = sorted(
    set((r['month'], r['year']) for r in form_records),
    key=lambda x: (int(x[1]), MONTHS.index(x[0])) if x[0] in MONTHS else (0,0)
)
MONTHLY_STATUS = {}
for mon, yr in all_months:
    recs = [r for r in form_records if r['month']==mon and r['year']==yr]
    vd = set(r['dist_div'].upper().strip() for r in recs if r['verified_by']=='EX EN DISTRIBUTION' and r['dist_div'])
    vt = set(r['test_div'].upper().strip() for r in recs if r['verified_by']=='EX EN TEST' and r['test_div'])
    vc = set(r['circle'].upper().strip() for r in recs if r['verified_by']=='CIRCLE OFFICE - SE' and r['circle'])
    result = []
    for m in MASTER_ALL:
        d = m['division'].upper().strip()
        if   m['type']=='DIST':   status = 'YES' if d in vd else 'NO'
        elif m['type']=='TEST':   status = 'YES' if d in vt else 'NO'
        else:                     status = 'YES' if d in vc else 'NO'
        result.append({'zone':m['zone'],'circle':m['circle'],
                       'type':m['type'],'division':m['division'],'status':status})
    MONTHLY_STATUS[f"{mon}_{yr}"] = result
    yes = sum(1 for r in result if r['status']=='YES')
    print(f"  {mon:12} {yr}: {yes}/{len(result)} ({round(yes/len(result)*100)}%)")

# mini form records for JS (compact array)
FORM_RECORDS = [[r['zone'],r['verified_by'],r['test_div'],r['dist_div'],
                 r['circle'],r['month'],r['year'],r['letter_no']] for r in form_records]

# ── 5. 5KW_INVOICE_DATA ───────────────────────────────────
print("\n[5KW] Reading 5KW_INVOICE_DATA...")
RECORDS_5KW = []
ws5 = wb['5KW_INVOICE_DATA']
for row in ws5.iter_rows(min_row=3, values_only=True):
    if not row[0]: continue
    if str(row[0]).strip().lower() in SKIP_ZONES: continue
    z = zone(row[0])
    if not z: continue
    if not isinstance(row[5], datetime): continue
    dt = row[5]
    s30 = str(row[16] or '').strip()
    # 30% amount rule:
    # Under Verification / Re-Verification / Annexure → col[17] (basic 30%)
    # Submit to HV Cell / ERP / CPC / Paid           → col[18] (actual 30% after verification)
    _upper = s30.upper()
    _use_actual = any(x in _upper for x in ['SUBMIT TO HV','ERP','CPC','PAID'])
    RECORDS_5KW.append({
        'zone':      z,
        'month':     MONTHS[dt.month - 1],
        'year':      str(dt.year),
        'invoice':   str(row[3] or '').strip(),
        'basic':     sf(row[11]),
        'gst':       sf(row[12]),
        'total':     sf(row[13]),
        'status_70': str(row[14] or '').strip(),
        'amt_70':    sf(row[15]),
        'status_30': s30,
        'amt_30':    sf(row[18]) if _use_actual else sf(row[17]),
        'penalty':   sf(row[19]),
        'from_zone': str(row[20] or '').strip(),
        'submit':    str(row[21] or '').strip(),
    })
paid_5kw   = sum(1 for r in RECORDS_5KW if r['status_30']=='PAID')
pending_5kw = len(RECORDS_5KW) - paid_5kw
months_5kw  = len(set((r['month'],r['year']) for r in RECORDS_5KW))
print(f"[5KW] Records: {len(RECORDS_5KW)} | Months: {months_5kw} | Paid: {paid_5kw} | Pending: {pending_5kw}")

# ── 6. 10KW_DATA ──────────────────────────────────────────
print("\n[10KW] Reading 10KW_DATA...")
RECORDS_10KW = []
ws10 = wb['10KW_DATA']
for row in ws10.iter_rows(min_row=3, values_only=True):
    if not row[0]: continue
    if not isinstance(row[1], datetime): continue
    dt = row[1]
    status = str(row[2] or '').strip()
    RECORDS_10KW.append({
        'month':  MONTHS[dt.month - 1],
        'year':   str(dt.year),
        'status': status,
        'amount': sf(row[3]),
        'paid':   status.upper() == 'PAID',
    })
paid_10kw = sum(1 for r in RECORDS_10KW if r['paid'])
pend_amt  = sum(r['amount'] for r in RECORDS_10KW if not r['paid'])
print(f"[10KW] Records: {len(RECORDS_10KW)} | Paid: {paid_10kw} | Pending: {len(RECORDS_10KW)-paid_10kw} | Pending Amt: ₹{pend_amt:,.0f}")

# ── 7. CLUSTER-4_DATA ─────────────────────────────────────
print("\n[C4] Reading CLUSTER-4_DATA...")
RECORDS_C4 = []
ws4 = wb['CLUSTER-4_DATA']
for row in ws4.iter_rows(min_row=3, values_only=True):
    if not row[0]: continue
    if str(row[0]).strip().lower() in SKIP_ZONES: continue
    z = zone(row[0])
    if not z: continue
    if not isinstance(row[4], datetime): continue
    dt = row[4]
    RECORDS_C4.append({
        'zone':      z,
        'month':     MONTHS[dt.month - 1],
        'year':      str(dt.year),
        'invoice':   str(row[2] or '').strip(),
        'probe':     sf(row[5]),
        'ocr':       sf(row[6]),
        'manual':    sf(row[7]),
        'basic':     sf(row[8]),
        'gst':       sf(row[9]),
        'total':     sf(row[10]),
        'status_70': str(row[11] or '').strip(),
        'amt_70':    sf(row[12]),
        'status_30': str(row[13] or '').strip(),
        'amt_30':    sf(row[14]),
    })
months_c4 = len(set((r['month'],r['year']) for r in RECORDS_C4))
zones_c4  = sorted(set(r['zone'] for r in RECORDS_C4))
print(f"[C4] Records: {len(RECORDS_C4)} | Months: {months_c4} | Zones: {zones_c4}")

# ── 8. Write data_v4.js ───────────────────────────────────
print("\n[Build] Writing data_v4.js...")
ts = datetime.now().strftime('%d %b %Y %H:%M')
data_js = (
    f"// Auto-updated: {ts}\n"
    f"// Form:{len(form_records)} | 5KW:{len(RECORDS_5KW)}recs/{months_5kw}mo"
    f" | 10KW:{len(RECORDS_10KW)}mo | C4:{len(RECORDS_C4)}recs/{months_c4}mo\n"
    f"const MASTER_ALL={json.dumps(MASTER_ALL,   separators=(',',':'))};\n"
    f"const MONTHLY_STATUS={json.dumps(MONTHLY_STATUS, separators=(',',':'))};\n"
    f"const FORM_RECORDS={json.dumps(FORM_RECORDS,   separators=(',',':'))};\n"
    f"var RECORDS_5KW={json.dumps(RECORDS_5KW,    separators=(',',':'))};\n"
    f"var RECORDS_10KW={json.dumps(RECORDS_10KW,   separators=(',',':'))};\n"
    f"var RECORDS_C4={json.dumps(RECORDS_C4,     separators=(',',':'))};\n"
)
with open('data_v4.js', 'w', encoding='utf-8') as f:
    f.write(data_js)
print(f"[Build] data_v4.js: {os.path.getsize('data_v4.js')//1024} KB")

# ── 9. Rebuild index.html from template ──────────────────────────────────────
print("[Build] Rebuilding index.html from index_template.html...")
with open('index_template.html', encoding='utf-8') as f:
    template = f.read()
index_html = template.replace('__DATA_PLACEHOLDER__', data_js)
with open('index.html', 'w', encoding='utf-8') as f:
    f.write(index_html)
print(f"[Build] index.html: {os.path.getsize('index.html')//1024} KB")

print(f"\n{'='*55}")
print(f"  ✅ Done! {ts}")
print(f"  Verification: {len(MONTHLY_STATUS)} months")
print(f"  5KW:  {len(RECORDS_5KW):3d} records | {months_5kw} months")
print(f"  10KW: {len(RECORDS_10KW):3d} records | {paid_10kw} paid | {len(RECORDS_10KW)-paid_10kw} pending")
print(f"  C4:   {len(RECORDS_C4):3d} records | {months_c4} months | {zones_c4}")
print(f"{'='*55}\n")
